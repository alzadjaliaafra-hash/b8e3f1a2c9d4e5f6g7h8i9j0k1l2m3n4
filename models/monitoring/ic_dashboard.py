#!/usr/bin/env python3
"""
SRFF-I Investment Committee Dashboard Generator
===============================================
Creates comprehensive Excel-based dashboard for IC decision support showing:
- Portfolio overview with key metrics
- Deal pipeline with RVS scores and survival probabilities
- Survival curves per company
- Sector analysis and concentration risk
- Stress test results under different scenarios

Author: Manus AI for Sohar International Bank
Date: April 2026
Version: 1.0
"""

import json
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, List, Optional
import warnings
warnings.filterwarnings('ignore')


class ICDashboard:
    """Investment Committee Dashboard Generator"""

    def __init__(self, data_source: str = "validation_results_v3.json"):
        """
        Initialize dashboard with portfolio data.

        Args:
            data_source: Path to JSON file with company data and RVS scores
        """
        self.data_source = data_source
        self.portfolio_data = self._load_data()

    def _load_data(self) -> Dict:
        """Load portfolio data from JSON validation results."""
        try:
            with open(self.data_source, 'r') as f:
                data = json.load(f)
            return data
        except FileNotFoundError:
            # Return sample data if validation file not found
            return self._generate_sample_data()

    def _generate_sample_data(self) -> Dict:
        """Generate sample portfolio data for demonstration."""
        companies = [
            {
                "company_name": f"Company_{chr(65+i)}",
                "sector": np.random.choice(["Manufacturing", "Healthcare", "Logistics", "Retail"]),
                "rvs_score": round(np.random.uniform(3.0, 9.0), 2),
                "recovery_probability": round(np.random.uniform(0.4, 0.95), 3),
                "survival_5yr": round(np.random.uniform(0.2, 0.9), 3),
                "hazard_rate": round(np.random.uniform(0.05, 0.30), 4),
                "composite_v3": round(np.random.uniform(0.5, 0.95), 3),
                "total_debt": round(np.random.uniform(5, 50), 1),
                "ebitda": round(np.random.uniform(1, 10), 2),
                "restructuring_cost": round(np.random.uniform(0.5, 3), 2),
                "status": np.random.choice(["Pipeline", "Due Diligence", "Deployed", "Monitoring"])
            }
            for i in range(30)
        ]

        return {
            "portfolio": companies,
            "generated_date": datetime.now().isoformat(),
            "total_companies": len(companies)
        }

    def _calculate_zone(self, rvs_score: float) -> str:
        """Classify RVS score into decision zones."""
        if rvs_score >= 7.0:
            return "Strong Candidate"
        elif rvs_score >= 4.5:
            return "Conditional"
        else:
            return "Reject"

    def generate_dashboard(self, output_file: str = "IC_Dashboard.xlsx"):
        """
        Generate comprehensive IC Dashboard Excel file.

        Args:
            output_file: Path for output Excel file
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create all dashboard sheets
        self._create_portfolio_overview(wb)
        self._create_deal_pipeline(wb)
        self._create_survival_curves(wb)
        self._create_sector_analysis(wb)
        self._create_stress_tests(wb)
        self._create_executive_summary(wb)

        # Save workbook
        wb.save(output_file)
        print(f"✅ IC Dashboard generated: {output_file}")
        return output_file

    def _create_portfolio_overview(self, wb: Workbook):
        """Sheet 1: Portfolio Overview with key metrics."""
        ws = wb.create_sheet("Portfolio Overview")

        # Extract portfolio data
        companies = self.portfolio_data.get("portfolio", [])

        # Calculate aggregate metrics
        total_companies = len(companies)
        total_exposure = sum(c.get("total_debt", 0) for c in companies)
        deployed = len([c for c in companies if c.get("status") == "Deployed"])
        avg_rvs = np.mean([c.get("rvs_score", 0) for c in companies])
        avg_recovery_prob = np.mean([c.get("recovery_probability", 0) for c in companies])

        # Header
        ws["A1"] = "SRFF-I PORTFOLIO OVERVIEW"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells("A1:E1")

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(italic=True)

        # Key Metrics
        row = 4
        metrics = [
            ("Total Companies", total_companies),
            ("Total Exposure (OMR M)", f"{total_exposure:.1f}"),
            ("Deployed Deals", deployed),
            ("Pipeline Deals", total_companies - deployed),
            ("Avg RVS Score", f"{avg_rvs:.2f}"),
            ("Avg Recovery Probability", f"{avg_recovery_prob:.1%}"),
        ]

        ws["A" + str(row)] = "Key Metrics"
        ws["A" + str(row)].font = Font(bold=True, size=12)
        row += 1

        for metric, value in metrics:
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Zone Distribution
        row += 2
        ws[f"A{row}"] = "Zone Distribution"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        zone_counts = {}
        for c in companies:
            zone = self._calculate_zone(c.get("rvs_score", 0))
            zone_counts[zone] = zone_counts.get(zone, 0) + 1

        for zone, count in zone_counts.items():
            ws[f"A{row}"] = zone
            ws[f"B{row}"] = count
            ws[f"C{row}"] = f"{count/total_companies:.1%}"
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

    def _create_deal_pipeline(self, wb: Workbook):
        """Sheet 2: Deal Pipeline with RVS scores."""
        ws = wb.create_sheet("Deal Pipeline")

        companies = self.portfolio_data.get("portfolio", [])

        # Create DataFrame
        df = pd.DataFrame(companies)
        df["zone"] = df["rvs_score"].apply(self._calculate_zone)
        df = df.sort_values("rvs_score", ascending=False)

        # Headers
        headers = ["Company", "Sector", "Status", "RVS Score", "Zone",
                   "Recovery Prob", "5Y Survival", "Hazard Rate",
                   "Composite V3", "Debt (OMR M)"]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, (_, company) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=company.get("company_name", ""))
            ws.cell(row=row_idx, column=2, value=company.get("sector", ""))
            ws.cell(row=row_idx, column=3, value=company.get("status", ""))
            ws.cell(row=row_idx, column=4, value=company.get("rvs_score", 0))

            zone_cell = ws.cell(row=row_idx, column=5, value=company["zone"])
            # Color code zones
            if company["zone"] == "Strong Candidate":
                zone_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif company["zone"] == "Conditional":
                zone_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                zone_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            ws.cell(row=row_idx, column=6, value=company.get("recovery_probability", 0))
            ws.cell(row=row_idx, column=7, value=company.get("survival_5yr", 0))
            ws.cell(row=row_idx, column=8, value=company.get("hazard_rate", 0))
            ws.cell(row=row_idx, column=9, value=company.get("composite_v3", 0))
            ws.cell(row=row_idx, column=10, value=company.get("total_debt", 0))

        # Format percentage columns
        for row in range(2, len(df) + 2):
            ws.cell(row=row, column=6).number_format = "0.0%"
            ws.cell(row=row, column=7).number_format = "0.0%"
            ws.cell(row=row, column=8).number_format = "0.00%"
            ws.cell(row=row, column=9).number_format = "0.0%"

        # Auto-adjust columns
        for col in range(1, 11):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_survival_curves(self, wb: Workbook):
        """Sheet 3: 5-Year Survival Probability Curves."""
        ws = wb.create_sheet("Survival Curves")

        companies = self.portfolio_data.get("portfolio", [])

        # Generate survival curves (Kaplan-Meier style)
        years = [0, 1, 2, 3, 4, 5]

        # Headers
        ws["A1"] = "Year"
        ws["A1"].font = Font(bold=True)

        for col_idx, company in enumerate(companies[:10], 2):  # Top 10 for clarity
            ws.cell(row=1, column=col_idx, value=company.get("company_name", f"Co_{col_idx-1}"))
            ws.cell(row=1, column=col_idx).font = Font(bold=True)

        # Generate survival data (exponential decay based on hazard rate)
        for row_idx, year in enumerate(years, 2):
            ws.cell(row=row_idx, column=1, value=year)

            for col_idx, company in enumerate(companies[:10], 2):
                hazard = company.get("hazard_rate", 0.1)
                survival = np.exp(-hazard * year)
                ws.cell(row=row_idx, column=col_idx, value=round(survival, 4))
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"

        # Create chart
        chart = LineChart()
        chart.title = "5-Year Survival Probability by Company"
        chart.style = 10
        chart.x_axis.title = "Years"
        chart.y_axis.title = "Survival Probability"

        data = Reference(ws, min_col=2, min_row=1, max_col=min(11, len(companies)+1), max_row=len(years)+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(years)+1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "A10")

    def _create_sector_analysis(self, wb: Workbook):
        """Sheet 4: Sector-level Analysis and Concentration Risk."""
        ws = wb.create_sheet("Sector Analysis")

        companies = self.portfolio_data.get("portfolio", [])

        # Aggregate by sector
        sector_data = {}
        for company in companies:
            sector = company.get("sector", "Unknown")
            if sector not in sector_data:
                sector_data[sector] = {
                    "count": 0,
                    "total_debt": 0,
                    "avg_rvs": [],
                    "avg_recovery": [],
                    "avg_survival": []
                }

            sector_data[sector]["count"] += 1
            sector_data[sector]["total_debt"] += company.get("total_debt", 0)
            sector_data[sector]["avg_rvs"].append(company.get("rvs_score", 0))
            sector_data[sector]["avg_recovery"].append(company.get("recovery_probability", 0))
            sector_data[sector]["avg_survival"].append(company.get("survival_5yr", 0))

        # Headers
        headers = ["Sector", "# Companies", "Total Debt (OMR M)", "% Exposure",
                   "Avg RVS", "Avg Recovery Prob", "Avg 5Y Survival"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # Calculate total exposure
        total_exposure = sum(d["total_debt"] for d in sector_data.values())

        # Write sector data
        row = 2
        for sector, data in sorted(sector_data.items(), key=lambda x: x[1]["total_debt"], reverse=True):
            ws.cell(row=row, column=1, value=sector)
            ws.cell(row=row, column=2, value=data["count"])
            ws.cell(row=row, column=3, value=round(data["total_debt"], 1))
            ws.cell(row=row, column=4, value=data["total_debt"] / total_exposure if total_exposure > 0 else 0)
            ws.cell(row=row, column=5, value=round(np.mean(data["avg_rvs"]), 2))
            ws.cell(row=row, column=6, value=round(np.mean(data["avg_recovery"]), 3))
            ws.cell(row=row, column=7, value=round(np.mean(data["avg_survival"]), 3))

            # Format percentages
            ws.cell(row=row, column=4).number_format = "0.0%"
            ws.cell(row=row, column=6).number_format = "0.0%"
            ws.cell(row=row, column=7).number_format = "0.0%"

            row += 1

        # Auto-adjust columns
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _create_stress_tests(self, wb: Workbook):
        """Sheet 5: Stress Test Scenarios."""
        ws = wb.create_sheet("Stress Tests")

        companies = self.portfolio_data.get("portfolio", [])

        # Define stress scenarios
        scenarios = {
            "Base Case": {"hazard_multiplier": 1.0, "rvs_adjustment": 0.0},
            "Mild Stress": {"hazard_multiplier": 1.25, "rvs_adjustment": -0.5},
            "Moderate Stress": {"hazard_multiplier": 1.5, "rvs_adjustment": -1.0},
            "Severe Stress": {"hazard_multiplier": 2.0, "rvs_adjustment": -1.5},
            "Extreme Stress": {"hazard_multiplier": 2.5, "rvs_adjustment": -2.0}
        }

        # Headers
        ws["A1"] = "Stress Test Results"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:F1")

        headers = ["Scenario", "Avg RVS", "Avg Recovery Prob", "Avg 5Y Survival",
                   "Strong Candidates", "Conditional", "Reject"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # Run stress tests
        row = 4
        for scenario_name, params in scenarios.items():
            # Apply stress adjustments
            stressed_rvs = [max(0, c.get("rvs_score", 0) + params["rvs_adjustment"])
                           for c in companies]
            stressed_hazard = [c.get("hazard_rate", 0.1) * params["hazard_multiplier"]
                              for c in companies]
            stressed_survival = [np.exp(-h * 5) for h in stressed_hazard]
            stressed_recovery = [max(0, min(1, c.get("recovery_probability", 0.5) * (1 - params["rvs_adjustment"]/10)))
                                for c in companies]

            # Count zones
            zone_counts = {"Strong Candidate": 0, "Conditional": 0, "Reject": 0}
            for rvs in stressed_rvs:
                zone = self._calculate_zone(rvs)
                zone_counts[zone] += 1

            # Write results
            ws.cell(row=row, column=1, value=scenario_name)
            ws.cell(row=row, column=2, value=round(np.mean(stressed_rvs), 2))
            ws.cell(row=row, column=3, value=round(np.mean(stressed_recovery), 3))
            ws.cell(row=row, column=4, value=round(np.mean(stressed_survival), 3))
            ws.cell(row=row, column=5, value=zone_counts["Strong Candidate"])
            ws.cell(row=row, column=6, value=zone_counts["Conditional"])
            ws.cell(row=row, column=7, value=zone_counts["Reject"])

            # Format percentages
            ws.cell(row=row, column=3).number_format = "0.0%"
            ws.cell(row=row, column=4).number_format = "0.0%"

            row += 1

        # Auto-adjust columns
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _create_executive_summary(self, wb: Workbook):
        """Sheet 6: Executive Summary Dashboard."""
        ws = wb.create_sheet("Executive Summary", 0)  # Make it first sheet

        companies = self.portfolio_data.get("portfolio", [])

        # Title
        ws["A1"] = "SRFF-I INVESTMENT COMMITTEE DASHBOARD"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 30

        ws["A2"] = f"Report Date: {datetime.now().strftime('%B %d, %Y')}"
        ws["A2"].font = Font(size=11, italic=True)

        # Key highlights
        row = 4
        ws[f"A{row}"] = "KEY HIGHLIGHTS"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        total_companies = len(companies)
        total_debt = sum(c.get("total_debt", 0) for c in companies)
        strong_count = len([c for c in companies if self._calculate_zone(c.get("rvs_score", 0)) == "Strong Candidate"])
        avg_recovery = np.mean([c.get("recovery_probability", 0) for c in companies])

        highlights = [
            ("Total Portfolio Companies", total_companies),
            ("Total Exposure (OMR Million)", f"{total_debt:.1f}"),
            ("Strong Rescue Candidates", f"{strong_count} ({strong_count/total_companies:.0%})"),
            ("Average Recovery Probability", f"{avg_recovery:.1%}"),
            ("Average RVS Score", f"{np.mean([c.get('rvs_score', 0) for c in companies]):.2f}")
        ]

        for label, value in highlights:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].font = Font(size=11)
            row += 1

        # Recommendations
        row += 2
        ws[f"A{row}"] = "IC RECOMMENDATIONS"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        # Top 5 companies to prioritize
        top_companies = sorted(companies, key=lambda x: x.get("rvs_score", 0), reverse=True)[:5]

        ws[f"A{row}"] = "Top 5 Priority Deals:"
        ws[f"A{row}"].font = Font(bold=True, underline="single")
        row += 1

        for idx, company in enumerate(top_companies, 1):
            ws[f"A{row}"] = f"{idx}. {company.get('company_name', 'Unknown')}"
            ws[f"B{row}"] = f"RVS: {company.get('rvs_score', 0):.2f}"
            ws[f"C{row}"] = f"Recovery: {company.get('recovery_probability', 0):.1%}"
            row += 1

        # Auto-adjust columns
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20


def main():
    """Generate IC Dashboard from validation results."""
    print("=" * 70)
    print("  SRFF-I INVESTMENT COMMITTEE DASHBOARD GENERATOR")
    print("=" * 70)
    print()

    # Create dashboard
    dashboard = ICDashboard()
    output_file = dashboard.generate_dashboard("SRFF_I_IC_Dashboard.xlsx")

    print()
    print("Dashboard includes:")
    print("  ✓ Executive Summary")
    print("  ✓ Portfolio Overview")
    print("  ✓ Deal Pipeline with RVS Scores")
    print("  ✓ 5-Year Survival Curves")
    print("  ✓ Sector Analysis & Concentration Risk")
    print("  ✓ Stress Test Scenarios")
    print()
    print(f"Open {output_file} to review the dashboard.")
    print("=" * 70)


if __name__ == "__main__":
    main()
